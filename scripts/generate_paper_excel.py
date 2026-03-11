#!/usr/bin/env python3
"""Generate separate Excel workbooks with SOMA-conformant data from two PM2.5 papers.

Each workbook has multiple tabs matching the LinkML schema classes (as in project/excel/soma.xlsx),
with data extracted from the YAML test files in tests/data/valid/.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def _add_row(ws, row, data, max_col):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def _auto_width(ws, max_col, min_w=12, max_w=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(
            max(min_w, max(len(str(c.value or "")) for c in ws[letter])),
            max_w,
        )


def _make_sheet(wb, name, headers, rows, tab_color=None):
    """Create a styled sheet with headers and data rows."""
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    ncol = len(headers)
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    _style_header(ws, 1, ncol)
    for i, row_data in enumerate(rows, 2):
        _add_row(ws, i, row_data, ncol)
    _auto_width(ws, ncol)
    return ws


def _make_metadata_sheet(wb, meta_rows):
    """Create the Metadata tab."""
    ws = wb.create_sheet("Metadata")
    ws.sheet_properties.tabColor = "4472C4"
    for i, row in enumerate(meta_rows, 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if i == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80


# =============================================================================
# SCHEMA HEADERS — matching project/excel/soma.xlsx tab columns
# =============================================================================

PROTOCOL_HEADERS = [
    "id", "name", "description", "protocol_type", "protocol_version",
    "equipment_required",
]

STAINING_PROTOCOL_HEADERS = [
    "id", "name", "description", "staining_type", "antibodies_used",
    "detection_method", "normalization_method", "fixation_method", "counterstain",
    "protocol_type", "equipment_required",
]

MOLECULAR_ASSAY_PROTOCOL_HEADERS = [
    "id", "name", "description", "detection_method", "normalization_method",
    "antibodies_used", "primer_sequences", "reference_gene",
    "protocol_type", "equipment_required",
]

SPIROMETRY_PROTOCOL_HEADERS = [
    "id", "name", "description", "spirometry_standard", "bronchodilator_agent",
    "bronchodilator_dose", "plethysmography_method",
    "protocol_type", "equipment_required",
]

EXPOSURE_CONDITION_HEADERS = [
    "id", "name", "exposure_agent", "exposure_agent_id",
    "exposure_concentration_value", "exposure_concentration_unit",
    "exposure_duration_value", "exposure_duration_unit",
]

KEY_EVENT_HEADERS = [
    "id", "name", "description", "biological_action",
    "level_of_biological_organization",
]

CELLULAR_SYSTEM_HEADERS = [
    "id", "name", "description", "subject_type",
    "cell_line", "cell_line_id", "primary_cell", "cell_type", "cell_type_id",
    "anatomical_origin", "model_species", "model_species_id",
    "cell_culture_growth_mode", "substrate_type",
    "days_at_differentiation", "donor_info",
]

IN_VIVO_SUBJECT_HEADERS = [
    "id", "name", "description", "subject_type",
    "model_species", "model_species_id",
    "age_value", "age_unit", "sex",
    "subject_characteristics", "disease_state",
    "sample_type", "collection_site",
]

CFTR_ASSAY_HEADERS = [
    "id", "name", "description",
    "stimulation_agent", "inhibitor_used",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

CFTR_OUTPUT_HEADERS = [
    "id", "name", "description",
    "cftr_chloride_secretion_value", "cftr_chloride_secretion_unit",
    "cftr_forskolin_response_value", "cftr_forskolin_response_unit",
    "inhibitor_sensitive_current_value", "inhibitor_sensitive_current_unit",
    "source_assay",
]

GENE_EXPRESSION_ASSAY_HEADERS = [
    "id", "name", "description",
    "target_gene", "gene_expression_method", "normalization_reference",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

GENE_EXPRESSION_OUTPUT_HEADERS = [
    "id", "name", "description",
    "mrna_level_value", "mrna_level_unit",
    "protein_level_value", "protein_level_unit",
    "percentage_positive_cells_value", "percentage_positive_cells_unit",
    "source_assay",
]

GOBLET_CELL_ASSAY_HEADERS = [
    "id", "name", "description",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

GOBLET_CELL_OUTPUT_HEADERS = [
    "id", "name", "description",
    "goblet_cell_percentage_value", "goblet_cell_percentage_unit",
    "muc5ac_mrna_expression_value", "muc5ac_mrna_expression_unit",
    "muc5ac_protein_expression_value", "muc5ac_protein_expression_unit",
    "muc5b_mrna_expression_value", "muc5b_mrna_expression_unit",
    "muc5b_protein_expression_value", "muc5b_protein_expression_unit",
    "mucin_secretion_rate_value", "mucin_secretion_rate_unit",
    "source_assay",
]

BALF_SPUTUM_ASSAY_HEADERS = [
    "id", "name", "description",
    "target_cell_type",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

BALF_SPUTUM_OUTPUT_HEADERS = [
    "id", "name", "description",
    "il6_concentration_value", "il6_concentration_unit",
    "source_assay",
]

LUNG_FUNCTION_ASSAY_HEADERS = [
    "id", "name", "description",
    "reference_dataset",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

LUNG_FUNCTION_OUTPUT_HEADERS = [
    "id", "name", "description",
    "lung_resistance_value", "lung_resistance_unit",
    "source_assay",
]

FOXJ_ASSAY_HEADERS = [
    "id", "name", "description",
    "informs_on_key_event", "study_subject",
    "has_exposure_condition", "follows_protocols",
    "assay_date",
]

FOXJ_OUTPUT_HEADERS = [
    "id", "name", "description",
    "foxj1_mrna_expression_value", "foxj1_mrna_expression_unit",
    "foxj1_positive_cell_percentage_value", "foxj1_positive_cell_percentage_unit",
    "source_assay",
]


# =============================================================================
# Montgomery et al. 2020
# =============================================================================

def create_montgomery_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Metadata ──
    _make_metadata_sheet(wb, [
        ["Field", "Value"],
        ["Short Name", "Montgomery2020"],
        ["Title", "Genome-Wide Analysis Reveals Mucociliary Remodeling of the Nasal Airway Epithelium Induced by Urban PM2.5"],
        ["Authors", "Montgomery MT, Sajuthi SP, Cho SH, Everman JL, Rios CL, Goldfarbmuren KC, Jackson ND, Saef B, Cromie M, Eng C, Medina V, Elhawary JR, Oh SS, Rodriguez-Santana J, Vladar EK, Burchard EG, Seibold MA"],
        ["Journal", "Am J Respir Cell Mol Biol 2020;63(2):172-184"],
        ["DOI", "10.1165/rcmb.2019-0454OC"],
        ["PMCID", "PMC7397762"],
        ["", ""],
        ["In Vitro Model", "Primary human nasal AECs at ALI, 28-32 days differentiation"],
        ["PM2.5 Source", "3 California cities (Bakersfield/Fresno, Sacramento, Yuba City), 2011"],
        ["PM2.5 Extract", "Water-soluble (WE) and organic-soluble (OE) fractions"],
        ["Key Concentrations", "OE: 0.045, 0.45, 4.5 ug/cm2; WE: 3, 30 ug/cm2"],
        ["Acute Protocol", "2 stimulations 24h apart, harvest 24h after second"],
        ["Chronic Protocol", "5 stimulations over 5 days (high OE dose)"],
        ["Toxicity", "<1% by LDH assay at all concentrations"],
        ["", ""],
        ["SOMA Assay Types", "GeneExpressionAssay, GobletCellAssay, FoxJExpressionAssay"],
    ])

    # ── Protocol ──
    _make_sheet(wb, "Protocol", PROTOCOL_HEADERS, [
        ("PROTOCOL:montgomery-rnaseq-001", "Whole transcriptome RNA-seq protocol", "",
         "MolecularAssayProtocol", "",
         "Illumina NovaSeq 6000; KAPA mRNA HyperPrep kit (Roche)"),
        ("PROTOCOL:montgomery-qpcr-001", "qRT-PCR validation protocol", "",
         "MolecularAssayProtocol", "",
         "Real-time PCR system"),
        ("PROTOCOL:montgomery-wb-001", "Western blot for SPDEF protein", "",
         "MolecularAssayProtocol", "",
         "SDS-PAGE apparatus; Densitometry imaging system"),
        ("PROTOCOL:montgomery-if-001", "Immunofluorescence for MUC5AC/MUC5B/FOXJ1", "",
         "StainingProtocol", "",
         "Confocal microscope; Anti-MUC5AC; Anti-MUC5B; Anti-FOXJ1; Anti-acetylated alpha-tubulin; DAPI"),
        ("PROTOCOL:montgomery-elisa-001", "Colorimetric ELISA for mucin secretion", "",
         "MolecularAssayProtocol", "",
         "Plate reader (490 nm); Anti-MUC5AC ELISA; Anti-MUC5B ELISA"),
        ("PROTOCOL:montgomery-abpas-001", "Alcian blue-PAS staining for mucins", "",
         "StainingProtocol", "",
         "Histological staining reagents (AB-PAS); Light microscope"),
    ], tab_color="70AD47")

    # ── ExposureCondition ──
    _make_sheet(wb, "ExposureCondition", EXPOSURE_CONDITION_HEADERS, [
        ("EXPOSURE:montgomery-mod-oe", "PM2.5 organic extract 0.45 ug/cm2 (moderate dose)",
         "PM2.5", "CHEBI:74481", "0.45", "ug/cm2 (UO:0000274)", "48", "hour (UO:0000032)"),
        ("EXPOSURE:montgomery-high-oe", "PM2.5 organic extract 4.5 ug/cm2 (high dose)",
         "PM2.5", "CHEBI:74481", "4.5", "ug/cm2 (UO:0000274)", "48", "hour (UO:0000032)"),
        ("EXPOSURE:montgomery-chronic-oe", "PM2.5 organic extract chronic 5-day stimulation (high OE dose)",
         "PM2.5", "CHEBI:74481", "4.5", "ug/cm2 (UO:0000274)", "120", "hour (UO:0000032)"),
    ], tab_color="FFC000")

    # ── KeyEvent ──
    _make_sheet(wb, "KeyEvent", KEY_EVENT_HEADERS, [
        ("KE:mie-ahr-activation", "AhR pathway activation", "CYP1A1 induction by PAH components of PM2.5", "activated", "molecular"),
        ("KE:ke-il1-inflammation", "IL-1 inflammatory program activation", "IL1A/IL1B hub cytokine network driving epithelial inflammation", "increased", "molecular"),
        ("KE:ke2-goblet-hyperplasia", "Goblet cell hyperplasia", "MUC5AC upregulation, mucus secretory cell enrichment", "increased", "cellular"),
        ("KE:ke-club-cell-loss", "Club cell marker loss", "SCGB1A1 downregulation indicating loss of club cell identity", "decreased", "cellular"),
        ("KE:ke-altered-ciliogenesis", "Altered ciliogenesis", "FOXJ1 downregulation and loss of ciliated cells", "decreased", "cellular"),
    ], tab_color="ED7D31")

    # ── CellularSystem ──
    _make_sheet(wb, "CellularSystem", CELLULAR_SYSTEM_HEADERS, [
        ("soma:montgomery-culture-001", "Primary nasal AEC ALI culture - GALA II donors",
         "Acute stimulation model (12 donors)", "CellularSystem",
         "", "", "nasal epithelial cell", "nasal epithelial cell", "CL:0002603",
         "nasal cavity (UBERON:0001707)", "Homo sapiens", "NCBITaxon:9606",
         "air_liquid_interface", "transwell_insert", 28,
         "12 donors (6 healthy, 6 asthmatic) from GALA II childhood asthma study"),
        ("soma:montgomery-culture-high", "Primary nasal AEC ALI culture - high OE dose",
         "High dose acute model (5 donors)", "CellularSystem",
         "", "", "nasal epithelial cell", "nasal epithelial cell", "CL:0002603",
         "nasal cavity (UBERON:0001707)", "Homo sapiens", "NCBITaxon:9606",
         "air_liquid_interface", "transwell_insert", 28,
         "5 donors from GALA II study"),
        ("soma:montgomery-culture-chronic", "Primary nasal AEC ALI culture - chronic stimulation",
         "5-day chronic stimulation model (4-5 donors)", "CellularSystem",
         "", "", "nasal epithelial cell", "nasal epithelial cell", "CL:0002603",
         "nasal cavity (UBERON:0001707)", "Homo sapiens", "NCBITaxon:9606",
         "air_liquid_interface", "transwell_insert", 28,
         "4-5 donors from GALA II study"),
    ], tab_color="4472C4")

    # ── GeneExpressionAssay ──
    _make_sheet(wb, "GeneExpressionAssay", GENE_EXPRESSION_ASSAY_HEADERS, [
        # Acute moderate OE (RNA-seq)
        ("GE:montgomery-cyp1a1-mod", "CYP1A1 mRNA - moderate OE",
         "Most highly upregulated gene; AhR pathway activation by PAH components",
         "PR:000005470", "RNA-seq (KAPA mRNA HyperPrep, NovaSeq 6000)", "DESeq2 median of ratios",
         "KE:mie-ahr-activation", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-mod-oe", "PROTOCOL:montgomery-rnaseq-001", "2020-01-01"),
        ("GE:montgomery-il1a-mod", "IL1A mRNA - moderate OE",
         "Hub cytokine driving broader epithelial inflammatory response",
         "PR:000001551", "RNA-seq", "DESeq2 median of ratios",
         "KE:ke-il1-inflammation", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-mod-oe", "PROTOCOL:montgomery-rnaseq-001", "2020-01-01"),
        ("GE:montgomery-il1b-mod", "IL1B mRNA - moderate OE",
         "Hub cytokine; drives mucus metaplasia via SPDEF/FOXA3/XBP1",
         "PR:000001554", "RNA-seq", "DESeq2 median of ratios",
         "KE:ke-il1-inflammation", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-mod-oe", "PROTOCOL:montgomery-rnaseq-001", "2020-01-01"),
        ("GE:montgomery-muc5ac-mod", "MUC5AC mRNA - moderate OE (RNA-seq)",
         "Mucus secretory cell enrichment Padj=1.19e-74",
         "PR:000011230", "RNA-seq", "DESeq2 median of ratios",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-mod-oe", "PROTOCOL:montgomery-rnaseq-001", "2020-01-01"),
        ("GE:montgomery-muc2-mod", "MUC2 mRNA - moderate OE (RNA-seq)",
         "Mucin gene upregulated at moderate dose",
         "PR:000011228", "RNA-seq", "DESeq2 median of ratios",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-mod-oe", "PROTOCOL:montgomery-rnaseq-001", "2020-01-01"),
        # Chronic qPCR
        ("GE:montgomery-spdef-chronic", "SPDEF mRNA - chronic OE stimulation",
         "Master TF for mucus metaplasia; protein 1.5-fold by WB",
         "PR:000014070", "qRT-PCR", "housekeeping gene",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe", "PROTOCOL:montgomery-qpcr-001; PROTOCOL:montgomery-wb-001", "2020-01-01"),
        ("GE:montgomery-scgb1a1-chronic", "SCGB1A1 mRNA - chronic OE stimulation",
         "Club cell marker loss after chronic PM2.5 OE; LFC=-1.51 (p=2.94e-2)",
         "PR:000014857", "qRT-PCR", "housekeeping gene",
         "KE:ke-club-cell-loss", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe", "PROTOCOL:montgomery-qpcr-001", "2020-01-01"),
    ], tab_color="5B9BD5")

    # ── GeneExpressionOutput ──
    _make_sheet(wb, "GeneExpressionOutput", GENE_EXPRESSION_OUTPUT_HEADERS, [
        ("GE:montgomery-cyp1a1-mod-output", "CYP1A1 moderate OE measurement", "",
         "6.21", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-cyp1a1-mod"),
        ("GE:montgomery-il1a-mod-output", "IL1A moderate OE measurement", "",
         "0.91", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-il1a-mod"),
        ("GE:montgomery-il1b-mod-output", "IL1B moderate OE measurement", "",
         "1.22", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-il1b-mod"),
        ("GE:montgomery-muc5ac-mod-output", "MUC5AC moderate OE RNA-seq measurement", "",
         "0.90", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-muc5ac-mod"),
        ("GE:montgomery-muc2-mod-output", "MUC2 moderate OE RNA-seq measurement", "",
         "1.34", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-muc2-mod"),
        ("GE:montgomery-spdef-chronic-output", "SPDEF chronic OE measurement", "",
         "0.84", "log2 fold change (UO:0000193)", "1.5", "fold change (UO:0000193)", "", "",
         "GE:montgomery-spdef-chronic"),
        ("GE:montgomery-scgb1a1-chronic-output", "SCGB1A1 chronic OE measurement", "",
         "-1.51", "log2 fold change (UO:0000193)", "", "", "", "", "GE:montgomery-scgb1a1-chronic"),
    ], tab_color="5B9BD5")

    # ── GobletCellAssay ──
    _make_sheet(wb, "GobletCellAssay", GOBLET_CELL_ASSAY_HEADERS, [
        ("GCM:montgomery-muc5ac-chronic", "MUC5AC expression and secretion - chronic OE",
         "Chronic 5-day PM2.5 OE. MUC5AC mRNA LFC=1.72 (p=1.27e-2). MUC5AC+ cells FC=2.88 (p=3.19e-2). MUC5AC secretion FC=2.2 (p=1.54e-2).",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe",
         "PROTOCOL:montgomery-qpcr-001; PROTOCOL:montgomery-if-001; PROTOCOL:montgomery-elisa-001",
         "2020-01-01"),
        ("GCM:montgomery-muc5b-chronic", "MUC5B expression and secretion - chronic OE",
         "Chronic 5-day PM2.5 OE. MUC5B mRNA LFC=-0.72 (p=9.74e-2, trend). MUC5B+ cells FC=0.74 (p=8.84e-2). MUC5B secretion FC=0.81 (p=3.59e-3).",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe",
         "PROTOCOL:montgomery-qpcr-001; PROTOCOL:montgomery-if-001",
         "2020-01-01"),
        ("GCM:montgomery-abpas-chronic", "Mucin-positive cells (AB-PAS) - chronic OE",
         "AB-PAS staining. Average FC=2.01 in mucin-positive cells (p=6.52e-2). 3 of 4 donors showed increase.",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe",
         "PROTOCOL:montgomery-abpas-001",
         "2020-01-01"),
        ("GCM:montgomery-ratio-mod", "MUC5AC/MUC5B ratio shift - moderate OE dose",
         "MUC5AC/MUC5B ratio significantly increased at moderate OE dose (p=3.19e-2). MUC5AC LFC=0.90 with MUC5B trending down.",
         "KE:ke2-goblet-hyperplasia", "soma:montgomery-culture-001",
         "EXPOSURE:montgomery-high-oe",
         "PROTOCOL:montgomery-rnaseq-001",
         "2020-01-01"),
    ], tab_color="A9D18E")

    # ── GobletCellOutput ──
    _make_sheet(wb, "GobletCellOutput", GOBLET_CELL_OUTPUT_HEADERS, [
        ("GCM:montgomery-muc5ac-chronic-output", "MUC5AC chronic OE measurements", "",
         "", "",
         "1.72", "log2 fold change (UO:0000193)",
         "2.88", "fold change (UO:0000193)",
         "", "",
         "", "",
         "2.2", "fold change (UO:0000193)",
         "GCM:montgomery-muc5ac-chronic"),
        ("GCM:montgomery-muc5b-chronic-output", "MUC5B chronic OE measurements", "",
         "", "",
         "", "",
         "", "",
         "-0.72", "log2 fold change (UO:0000193)",
         "0.74", "fold change (UO:0000193)",
         "", "",
         "GCM:montgomery-muc5b-chronic"),
        ("GCM:montgomery-abpas-chronic-output", "AB-PAS chronic OE measurements", "",
         "2.01", "fold change (UO:0000193)",
         "", "", "", "", "", "", "", "",
         "", "",
         "GCM:montgomery-abpas-chronic"),
        ("GCM:montgomery-ratio-mod-output", "MUC5AC/MUC5B ratio moderate OE", "",
         "", "",
         "0.90", "log2 fold change (UO:0000193)",
         "", "",
         "-0.79", "log2 fold change (UO:0000193)",
         "", "",
         "", "",
         "GCM:montgomery-ratio-mod"),
    ], tab_color="A9D18E")

    # ── FoxJExpressionAssay ──
    _make_sheet(wb, "FoxJExpressionAssay", FOXJ_ASSAY_HEADERS, [
        ("FOXJ:montgomery-chronic", "FOXJ1 expression - chronic OE stimulation",
         "FOXJ1 downregulation after chronic 5-day PM2.5 OE. mRNA LFC=-0.36 (p=1.92e-2). FOXJ1+ nuclei decreased 75% (p=1.20e-2). Weaker acetylated alpha-tubulin.",
         "KE:ke-altered-ciliogenesis", "soma:montgomery-culture-chronic",
         "EXPOSURE:montgomery-chronic-oe",
         "PROTOCOL:montgomery-qpcr-001; PROTOCOL:montgomery-if-001",
         "2020-01-01"),
        ("FOXJ:montgomery-high-acute", "FOXJ1 and ciliogenesis TFs - high OE acute dose",
         "High OE: ciliary assembly module downregulated (p=4.20e-4). FOXJ1, MCIDAS, MYB, TP73, RFX2, RFX3 all decreased. Ciliated cell enrichment Padj=3.57e-118.",
         "KE:ke-altered-ciliogenesis", "soma:montgomery-culture-high",
         "EXPOSURE:montgomery-high-oe",
         "PROTOCOL:montgomery-rnaseq-001",
         "2020-01-01"),
    ], tab_color="BF8F00")

    # ── FoxJExpressionOutput ──
    _make_sheet(wb, "FoxJExpressionOutput", FOXJ_OUTPUT_HEADERS, [
        ("FOXJ:montgomery-chronic-output", "FOXJ1 chronic OE measurements", "",
         "-0.36", "log2 fold change (UO:0000193)",
         "25", "percent (UO:0000187)",
         "FOXJ:montgomery-chronic"),
        ("FOXJ:montgomery-high-acute-output", "FOXJ1 high OE measurements", "",
         "-0.36", "log2 fold change (UO:0000193)",
         "", "",
         "FOXJ:montgomery-high-acute"),
    ], tab_color="BF8F00")

    out = "/Users/SMoxon/Documents/src/soma/src/docs/Montgomery2020_PM25_Mucociliary_SOMA.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


# =============================================================================
# Liu et al. 2024
# =============================================================================

def create_liu_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Metadata ──
    _make_metadata_sheet(wb, [
        ["Field", "Value"],
        ["Short Name", "Liu2024"],
        ["Title", "PM2.5 Exposure Inhibits Transepithelial Anion Short-circuit Current by Downregulating P2Y2 Receptor/CFTR Pathway"],
        ["Authors", "Liu X, Li Z, Shan J, Wang F, Li Z, Luo S, Wu J"],
        ["Journal", "Int J Med Sci 2024;21(10):1929-1944"],
        ["DOI", "10.7150/ijms.96777"],
        ["Ethics", "GDREC2019347A; SYXK(Guangdong)2017-0178"],
        ["", ""],
        ["In Vitro Model", "Calu-3 human airway epithelial cells, submerged monolayer, transwell (Corning 3801)"],
        ["In Vitro Medium", "DMEM + 10% FBS + 1% pen/strep, 21-day post-confluency, 37C, 5% CO2"],
        ["In Vivo Model", "Male Balb/c mice (6-8 wk, 18-22g, SPF), OVA-induced asthma (n=50, 5 groups of 10)"],
        ["In Vivo Protocol", "OVA sensitization d1,7,14 (IP); OVA challenge d21-27 (inhaled 1%); treatments d21-27 (nasal)"],
        ["PM2.5 Source", "Guangzhou, China (Jul-Nov 2019), high-flow sampler at 1.05 m3/min"],
        ["In Vitro PM2.5", "100 ug/mL in serum-free DMEM/F12; 24h and 48h exposures"],
        ["Treatment Groups", "Control (PBS), ATP (10 uM), PM2.5 (100 ug/mL), ATP+PM2.5, Suramin (100 uM), PM2.5+Suramin"],
        ["In Vivo Treatments", "Saline, OVA alone, OVA+PM2.5 (100 ug/mL nasal), OVA+ATP (50 mg/kg nasal), OVA+Suramin (50 mg/kg nasal)"],
        ["", ""],
        ["SOMA Assay Types", "CFTRFunctionAssay, GeneExpressionAssay, GobletCellAssay, BALFSputumAssay, LungFunctionAssay"],
    ])

    # ── Protocol ──
    _make_sheet(wb, "Protocol", PROTOCOL_HEADERS, [
        ("PROTOCOL:liu-ussing-001", "Ussing chamber short-circuit current protocol",
         "Kingtech chamber, K-H solutions (mannitol luminal / glucose basal), 37C, 95% O2/5% CO2, electrode >60 uA",
         "ElectrophysiologyProtocol", "",
         "Ussing chamber (Kingtech, Beijing); Electrophysiology detection equipment"),
        ("PROTOCOL:liu-wb-001", "Western blot for P2Y2R and CFTR",
         "RIPA lysis, BCA, 10% SDS-PAGE (40ug/lane), PVDF, 5% skim milk 1.5h. Primary: P2Y2R (Abcam ab272891, 1:1000), CFTR (CST 78335S, 1:1000). Secondary: anti-rabbit IgG HRP (CST 7074S, 1:2500).",
         "MolecularAssayProtocol", "",
         "SDS-PAGE apparatus; PVDF membrane (Millipore); ChemiScope imaging (Clinx, Shanghai)"),
        ("PROTOCOL:liu-qpcr-001", "qRT-PCR for P2Y2R",
         "TRIzol RNA extraction, Takara cDNA synthesis. P2Y2R-F: CCTGAGAGGAGAAGCGCAG, P2Y2R-R: GAACTCTGCGGGAAACAGGA. GAPDH-F: AGATCCCTCCAAAATCAAGTGG, GAPDH-R: GGCAGAGATGATGACCCTTTT.",
         "MolecularAssayProtocol", "",
         "Real-time PCR system; TRIzol reagent (Sigma); Reverse transcription kit (Takara)"),
        ("PROTOCOL:liu-flowcyt-001", "Fluo-3/AM calcium measurement by flow cytometry",
         "Trypsin (no EDTA), Fluo-3/AM probe (Beyotime) 0.5uL, 37C 30min, FITC channel (Ex 488nm, Em 525nm)",
         "MolecularAssayProtocol", "",
         "BD Accuri C6 Plus flow cytometer; Fluo-3/AM probe (Beyotime); FlowJo v10.0"),
        ("PROTOCOL:liu-ihc-001", "Immunohistochemistry for P2Y2R and CFTR",
         "4% PFA fixation 24h, paraffin 5um sections. P2Y2R (1:500, Abcam), CFTR (1:1000, Abclone). Mouse anti-rabbit IgG + SP complex (ZSGB-BIO), 37C 40min. DAB 0.05%.",
         "StainingProtocol", "",
         "Microtome; DAB substrate (ZSGB-BIO, Beijing)"),
        ("PROTOCOL:liu-pas-001", "PAS staining for goblet cells",
         "Standard PAS staining on 5um paraffin sections. Quantification by percentage PAS-positive cells.",
         "StainingProtocol", "",
         "PAS staining reagents (Sigma); Light microscope"),
        ("PROTOCOL:liu-elisa-001", "ELISA for Th2 cytokines in BALF",
         "Mouse ELISA kits (Absin, Shanghai) for IL-4, IL-5, IL-13. Per manufacturer instructions.",
         "MolecularAssayProtocol", "",
         "ELISA plate reader; Mouse IL-4 ELISA kit (Absin); Mouse IL-5 ELISA kit (Absin); Mouse IL-13 ELISA kit (Absin)"),
        ("PROTOCOL:liu-lungfx-001", "Whole-body plethysmography with methacholine challenge",
         "BUXCO whole-body chamber (DSI, Minnesota). Methacholine aerosol at 0, 5, 10, 20 mg/mL. sRaw (cmH2O/mL/s). FLexi Vent.",
         "SpirometryProtocol", "",
         "BUXCO whole-body plethysmograph (DSI, Minnesota); Methacholine nebulizer"),
    ], tab_color="70AD47")

    # ── ExposureCondition ──
    _make_sheet(wb, "ExposureCondition", EXPOSURE_CONDITION_HEADERS, [
        ("EXPOSURE:liu-pm25-100ug-24h", "PM2.5 100 ug/mL for 24 hours",
         "PM2.5", "CHEBI:74481", "100", "ug/mL (UO:0000064)", "24", "hour (UO:0000032)"),
        ("EXPOSURE:liu-pm25-100ug-48h", "PM2.5 100 ug/mL for 48 hours",
         "PM2.5", "CHEBI:74481", "100", "ug/mL (UO:0000064)", "48", "hour (UO:0000032)"),
        ("EXPOSURE:liu-mouse-ova-pm25", "OVA sensitization + PM2.5 nasal instillation",
         "PM2.5", "CHEBI:74481", "100", "ug/mL (UO:0000064)", "7", "day (UO:0000033)"),
        ("EXPOSURE:liu-mouse-ova-atp", "OVA sensitization + ATP 50 mg/kg nasal instillation",
         "ATP", "CHEBI:15422", "50", "mg/kg (UO:0000308)", "7", "day (UO:0000033)"),
        ("EXPOSURE:liu-mouse-suramin", "OVA sensitization + Suramin 50 mg/kg nasal instillation",
         "suramin", "CHEBI:45653", "50", "mg/kg (UO:0000308)", "7", "day (UO:0000033)"),
    ], tab_color="FFC000")

    # ── KeyEvent ──
    _make_sheet(wb, "KeyEvent", KEY_EVENT_HEADERS, [
        ("KE:ke-decreased-cftr", "Decreased CFTR function",
         "PM2.5 downregulates CFTR via P2Y2R pathway; inhibits ATP-induced Isc", "decreased", "molecular"),
        ("KE:ke-decreased-p2y2r", "Decreased P2Y2R expression",
         "PM2.5 downregulates P2Y2R mRNA and protein in Calu-3 and mouse lung", "decreased", "molecular"),
        ("KE:ke2-goblet-hyperplasia", "Goblet cell hyperplasia",
         "PAS-positive cells increased in OVA+PM2.5 mouse lung", "increased", "cellular"),
        ("KE:ke-airway-inflammation", "Th2 airway inflammation",
         "BALF IL-4, IL-5, IL-13 elevated in OVA+PM2.5", "increased", "tissue"),
        ("KE:ao-decreased-lung-function", "Increased airway hyperresponsiveness",
         "sRaw elevated at methacholine challenge", "increased", "organ"),
    ], tab_color="ED7D31")

    # ── CellularSystem ──
    _make_sheet(wb, "CellularSystem", CELLULAR_SYSTEM_HEADERS, [
        ("soma:liu-calu3-001", "Calu-3 submerged monolayer on transwell",
         "In vitro model for CFTR/P2Y2R studies", "CellularSystem",
         "Calu-3", "CLO:0003679", "", "epithelial cell", "CL:0000066",
         "", "Homo sapiens", "NCBITaxon:9606",
         "adherent", "transwell_insert", 21, ""),
    ], tab_color="4472C4")

    # ── InVivoSubject ──
    _make_sheet(wb, "InVivoSubject", IN_VIVO_SUBJECT_HEADERS, [
        ("SUBJECT:liu-mouse-control", "Control Balb/c mouse",
         "Saline control group", "InVivoSubject",
         "Mus musculus", "NCBITaxon:10090",
         "7", "week (UO:0000034)", "male",
         "SPF grade, 18-22g, saline control group", "", "biopsy", ""),
        ("SUBJECT:liu-mouse-ova-pm25", "OVA+PM2.5 Balb/c mouse",
         "OVA-sensitized + PM2.5 treated", "InVivoSubject",
         "Mus musculus", "NCBITaxon:10090",
         "7", "week (UO:0000034)", "male",
         "SPF grade, 18-22g, OVA-sensitized + PM2.5 treated", "OVA-induced asthma", "biopsy", "lung tissue"),
        ("SUBJECT:liu-mouse-ova-atp", "OVA+ATP Balb/c mouse",
         "OVA-sensitized + ATP treated", "InVivoSubject",
         "Mus musculus", "NCBITaxon:10090",
         "7", "week (UO:0000034)", "male",
         "SPF grade, 18-22g, OVA-sensitized + ATP treated", "OVA-induced asthma", "biopsy", ""),
        ("SUBJECT:liu-mouse-ova-suramin", "OVA+Suramin Balb/c mouse",
         "OVA-sensitized + Suramin treated", "InVivoSubject",
         "Mus musculus", "NCBITaxon:10090",
         "7", "week (UO:0000034)", "male",
         "SPF grade, 18-22g", "OVA-induced asthma", "biopsy", ""),
    ], tab_color="4472C4")

    # ── CFTRFunctionAssay ──
    _make_sheet(wb, "CFTRFunctionAssay", CFTR_ASSAY_HEADERS, [
        ("CFTR:liu-control-24h", "CFTR Isc - Control + ATP 24h",
         "ATP-induced transepithelial short-circuit current in control Calu-3 cells at 24h",
         "ATP 10 uM", "",
         "KE:ke-decreased-cftr", "soma:liu-calu3-001",
         "", "PROTOCOL:liu-ussing-001", "2024-01-01"),
        ("CFTR:liu-pm25-24h", "CFTR Isc - PM2.5 + ATP 24h",
         "ATP-induced Isc after 24h PM2.5 (100 ug/mL). Significantly slower Isc increase (p<0.05). dIsc/dT decreased.",
         "ATP 10 uM", "",
         "KE:ke-decreased-cftr", "soma:liu-calu3-001",
         "EXPOSURE:liu-pm25-100ug-24h", "PROTOCOL:liu-ussing-001", "2024-01-01"),
        ("CFTR:liu-mouse-control", "CFTR Isc - Mouse trachea control",
         "ATP-induced Isc in control mouse tracheal tissue (8mm x 8mm). ATP 100 uM luminal.",
         "ATP 100 uM", "",
         "KE:ke-decreased-cftr", "SUBJECT:liu-mouse-control",
         "", "PROTOCOL:liu-ussing-001", "2024-01-01"),
        ("CFTR:liu-mouse-pm25", "CFTR Isc - Mouse trachea PM2.5",
         "ATP-induced Isc in PM2.5-exposed mouse trachea. dIsc lower vs control (p<0.05). dIsc/dT decreased (p<0.05).",
         "ATP 100 uM", "",
         "KE:ke-decreased-cftr", "SUBJECT:liu-mouse-ova-pm25",
         "EXPOSURE:liu-mouse-ova-pm25", "PROTOCOL:liu-ussing-001", "2024-01-01"),
    ], tab_color="C00000")

    # ── CFTRFunctionOutput ──
    _make_sheet(wb, "CFTRFunctionOutput", CFTR_OUTPUT_HEADERS, [
        ("CFTR:liu-control-24h-output", "Control Isc measurement", "",
         "1.0", "fold change (UO:0000193)", "", "", "", "", "CFTR:liu-control-24h"),
        ("CFTR:liu-pm25-24h-output", "PM2.5 Isc measurement", "",
         "0.6", "fold change (UO:0000193)", "", "", "", "", "CFTR:liu-pm25-24h"),
        ("CFTR:liu-mouse-control-output", "Mouse trachea control Isc", "",
         "1.0", "fold change (UO:0000193)", "", "", "", "", "CFTR:liu-mouse-control"),
        ("CFTR:liu-mouse-pm25-output", "Mouse trachea PM2.5 Isc", "",
         "0.55", "fold change (UO:0000193)", "", "", "", "", "CFTR:liu-mouse-pm25"),
    ], tab_color="C00000")

    # ── GeneExpressionAssay ──
    _make_sheet(wb, "GeneExpressionAssay", GENE_EXPRESSION_ASSAY_HEADERS, [
        # P2Y2R mRNA
        ("GE:liu-p2y2r-pm25-24h", "P2Y2R mRNA - PM2.5 24h",
         "P2Y2R mRNA downregulated after 24h PM2.5 in Calu-3. ATP group no significant difference vs control.",
         "PR:000001815", "qRT-PCR", "GAPDH",
         "KE:ke-decreased-p2y2r", "soma:liu-calu3-001",
         "EXPOSURE:liu-pm25-100ug-24h", "PROTOCOL:liu-qpcr-001", "2024-01-01"),
        ("GE:liu-p2y2r-pm25-48h", "P2Y2R mRNA - PM2.5 48h",
         "P2Y2R mRNA further downregulated after 48h PM2.5. ATP group significantly higher vs control at 48h (p<0.05).",
         "PR:000001815", "qRT-PCR", "GAPDH",
         "KE:ke-decreased-p2y2r", "soma:liu-calu3-001",
         "EXPOSURE:liu-pm25-100ug-48h", "PROTOCOL:liu-qpcr-001", "2024-01-01"),
        # CFTR protein (WB)
        ("GE:liu-cftr-protein-pm25-24h", "CFTR protein - PM2.5 24h (Western blot)",
         "CFTR protein significantly decreased after 24h PM2.5 (p<0.05). ATP+PM2.5 << ATP alone (p<0.01).",
         "PR:000003411", "Western blot", "GAPDH",
         "KE:ke-decreased-cftr", "soma:liu-calu3-001",
         "EXPOSURE:liu-pm25-100ug-24h", "PROTOCOL:liu-wb-001", "2024-01-01"),
        # P2Y2R protein (WB)
        ("GE:liu-p2y2r-protein-pm25-24h", "P2Y2R protein - PM2.5 24h (Western blot)",
         "P2Y2R protein significantly decreased after 24h PM2.5 (p<0.05). ATP+PM2.5 << ATP alone (p<0.01).",
         "PR:000001815", "Western blot", "GAPDH",
         "KE:ke-decreased-p2y2r", "soma:liu-calu3-001",
         "EXPOSURE:liu-pm25-100ug-24h", "PROTOCOL:liu-wb-001", "2024-01-01"),
        # P2Y2R lung tissue IHC (in vivo)
        ("GE:liu-p2y2r-ihc-ova-pm25", "P2Y2R protein - lung tissue IHC (OVA+PM2.5)",
         "P2Y2R IHC positive area dramatically decreased in OVA+PM2.5. OVA+ATP significantly increased vs OVA (p<0.05).",
         "PR:000001815", "immunohistochemistry", "tissue area",
         "KE:ke-decreased-p2y2r", "SUBJECT:liu-mouse-ova-pm25",
         "EXPOSURE:liu-mouse-ova-pm25", "PROTOCOL:liu-ihc-001", "2024-01-01"),
    ], tab_color="5B9BD5")

    # ── GeneExpressionOutput ──
    _make_sheet(wb, "GeneExpressionOutput", GENE_EXPRESSION_OUTPUT_HEADERS, [
        ("GE:liu-p2y2r-pm25-24h-output", "P2Y2R mRNA 24h PM2.5", "",
         "0.55", "fold change (UO:0000193)", "", "", "", "", "GE:liu-p2y2r-pm25-24h"),
        ("GE:liu-p2y2r-pm25-48h-output", "P2Y2R mRNA 48h PM2.5", "",
         "0.40", "fold change (UO:0000193)", "", "", "", "", "GE:liu-p2y2r-pm25-48h"),
        ("GE:liu-cftr-protein-pm25-24h-output", "CFTR protein 24h PM2.5", "",
         "", "", "0.50", "fold change (UO:0000193)", "", "", "GE:liu-cftr-protein-pm25-24h"),
        ("GE:liu-p2y2r-protein-pm25-24h-output", "P2Y2R protein 24h PM2.5", "",
         "", "", "0.45", "fold change (UO:0000193)", "", "", "GE:liu-p2y2r-protein-pm25-24h"),
        ("GE:liu-p2y2r-ihc-ova-pm25-output", "P2Y2R lung tissue IHC", "",
         "", "", "", "", "15", "percent (UO:0000187)", "GE:liu-p2y2r-ihc-ova-pm25"),
    ], tab_color="5B9BD5")

    # ── GobletCellAssay ──
    _make_sheet(wb, "GobletCellAssay", GOBLET_CELL_ASSAY_HEADERS, [
        ("GCM:liu-pas-ova-pm25", "Goblet cell hyperplasia - OVA+PM2.5 (PAS)",
         "PAS staining in mouse lung. OVA+PM2.5 evidently higher goblet cell % than OVA alone (p<0.001). Primarily near trachea/bronchus.",
         "KE:ke2-goblet-hyperplasia", "SUBJECT:liu-mouse-ova-pm25",
         "EXPOSURE:liu-mouse-ova-pm25", "PROTOCOL:liu-pas-001", "2024-01-01"),
        ("GCM:liu-pas-ova-atp", "Goblet cell rescue - OVA+ATP (PAS)",
         "PAS staining in OVA+ATP mouse lung. Goblet cell % lower than OVA alone (p<0.001). ATP partially rescues.",
         "KE:ke2-goblet-hyperplasia", "SUBJECT:liu-mouse-ova-atp",
         "EXPOSURE:liu-mouse-ova-atp", "PROTOCOL:liu-pas-001", "2024-01-01"),
    ], tab_color="A9D18E")

    # ── GobletCellOutput ──
    _make_sheet(wb, "GobletCellOutput", GOBLET_CELL_OUTPUT_HEADERS, [
        ("GCM:liu-pas-ova-pm25-output", "PAS goblet cell quantification", "",
         "45", "percent (UO:0000187)",
         "", "", "", "", "", "", "", "",
         "", "",
         "GCM:liu-pas-ova-pm25"),
        ("GCM:liu-pas-ova-atp-output", "PAS goblet cell OVA+ATP", "",
         "18", "percent (UO:0000187)",
         "", "", "", "", "", "", "", "",
         "", "",
         "GCM:liu-pas-ova-atp"),
    ], tab_color="A9D18E")

    # ── BALFSputumAssay ──
    _make_sheet(wb, "BALFSputumAssay", BALF_SPUTUM_ASSAY_HEADERS, [
        ("BALF:liu-ova-pm25", "BALF Th2 cytokines - OVA+PM2.5",
         "BALF from right lung (left bronchus ligated, lavaged 0.5mL saline x2). IL-4, IL-5, IL-13 all significantly increased in OVA+PM2.5 vs OVA alone (p<0.05).",
         "",
         "KE:ke-airway-inflammation", "SUBJECT:liu-mouse-ova-pm25",
         "EXPOSURE:liu-mouse-ova-pm25", "PROTOCOL:liu-elisa-001", "2024-01-01"),
    ], tab_color="7030A0")

    # ── BALFSputumOutput ──
    _make_sheet(wb, "BALFSputumOutput", BALF_SPUTUM_OUTPUT_HEADERS, [
        ("BALF:liu-ova-pm25-output", "BALF cytokine measurements OVA+PM2.5",
         "IL-4, IL-5, IL-13 all increased. Measured via ELISA (Absin, Shanghai).",
         "1.8", "fold change vs OVA (UO:0000193)",
         "BALF:liu-ova-pm25"),
    ], tab_color="7030A0")

    # ── LungFunctionAssay ──
    _make_sheet(wb, "LungFunctionAssay", LUNG_FUNCTION_ASSAY_HEADERS, [
        ("LF:liu-ova-pm25", "Airway resistance (sRaw) - OVA+PM2.5",
         "Whole-body plethysmography with MCh challenge. OVA+PM2.5 sRaw further increased vs OVA at all MCh concentrations.",
         "",
         "KE:ao-decreased-lung-function", "SUBJECT:liu-mouse-ova-pm25",
         "EXPOSURE:liu-mouse-ova-pm25", "PROTOCOL:liu-lungfx-001", "2024-01-01"),
        ("LF:liu-ova-suramin", "Airway resistance (sRaw) - OVA+Suramin",
         "OVA+Suramin sRaw at 20 mg/mL MCh significantly higher than OVA alone (p<0.05). P2Y2R antagonism worsens AHR.",
         "",
         "KE:ao-decreased-lung-function", "SUBJECT:liu-mouse-ova-suramin",
         "EXPOSURE:liu-mouse-suramin", "PROTOCOL:liu-lungfx-001", "2024-01-01"),
    ], tab_color="002060")

    # ── LungFunctionOutput ──
    _make_sheet(wb, "LungFunctionOutput", LUNG_FUNCTION_OUTPUT_HEADERS, [
        ("LF:liu-ova-pm25-output", "sRaw measurement OVA+PM2.5", "",
         "2.5", "fold change vs control (UO:0000193)", "LF:liu-ova-pm25"),
        ("LF:liu-ova-suramin-output", "sRaw measurement OVA+Suramin", "",
         "3.0", "fold change vs control (UO:0000193)", "LF:liu-ova-suramin"),
    ], tab_color="002060")

    out = "/Users/SMoxon/Documents/src/soma/src/docs/Liu2024_PM25_CFTR_SOMA.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    create_montgomery_workbook()
    create_liu_workbook()
