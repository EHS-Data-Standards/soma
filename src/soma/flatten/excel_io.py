"""Excel I/O for SOMA flat data.

Writer: flat rows -> styled .xlsx (one sheet per assay type)
Reader: .xlsx -> flat rows -> delegates to Unflattener
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel I/O. Install with: pip install openpyxl"
    )


# Styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def _get_column_order(rows: list[dict[str, Any]]) -> list[str]:
    """Get ordered column names from rows, preserving insertion order."""
    seen: dict[str, None] = {}
    for row in rows:
        for key in row:
            if key not in seen:
                seen[key] = None
    return list(seen.keys())


def _auto_width(ws: Any, min_width: int = 12, max_width: int = 40) -> None:
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def write_excel(
    sheets_data: dict[str, list[dict[str, Any]]],
    output_path: str | Path,
    descriptions: dict[str, dict[str, str]] | None = None,
) -> Path:
    """Write flat rows to a styled Excel workbook.

    Args:
        sheets_data: Dict mapping sheet name to list of flat row dicts
        output_path: Path for the output .xlsx file
        descriptions: Optional dict mapping sheet_name -> {column_name: description}
                      Used for header comments/tooltips

    Returns:
        Path to the written file
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheets_data.items():
        if not rows:
            ws = wb.create_sheet(title=sheet_name)
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit
        columns = _get_column_order(rows)

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                val = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = DATA_ALIGNMENT

        # Auto-size columns
        _auto_width(ws)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def write_template_excel(
    specs: dict[str, dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """Generate an empty Excel workbook template from flattening specs.

    Creates one sheet per assay type with headers derived from the spec.
    """
    sheets_data: dict[str, list[dict[str, Any]]] = {}

    for assay_class, spec in specs.items():
        sheet_name = spec.get("target_sheet", assay_class)
        # Build column list from spec
        columns = _columns_from_spec(spec)
        # Create one empty row to establish column order
        empty_row = {col: None for col in columns}
        sheets_data[sheet_name] = [empty_row]

    return write_excel(sheets_data, output_path)


def _columns_from_spec(spec: dict[str, Any]) -> list[str]:
    """Derive column names from a flattening spec."""
    rules = spec["flatten_rules"]
    columns = []

    # Scalars
    columns.extend(rules.get("scalars", []))

    # QuantityValue outputs
    for qv in rules.get("quantity_values", []):
        prefix = qv["prefix"]
        columns.extend([f"{prefix}_value", f"{prefix}_unit"])

    # Output scalars
    for os_spec in rules.get("output_scalars", []):
        columns.append(os_spec["column"])

    # Inlined objects
    for inl in rules.get("inlined_objects", []):
        for col_spec in inl["columns"]:
            columns.append(col_spec["column"])

    # Study subject common
    if "study_subject_common" in rules:
        common = rules["study_subject_common"]
        columns.append("subject_type")
        for name in common.get("scalars", []):
            columns.append(f"subject_{name}")
        for inl in common.get("inlined_objects", []):
            for col_spec in inl["columns"]:
                columns.append(col_spec["column"])

    # Polymorphic: add all variant columns
    for poly in rules.get("polymorphic", []):
        for _variant_name, variant in poly.get("variants", {}).items():
            for name in variant.get("scalars", []):
                if name not in columns:
                    columns.append(name)
            for qv in variant.get("quantity_values", []):
                prefix = qv["prefix"]
                for col in [f"{prefix}_value", f"{prefix}_unit"]:
                    if col not in columns:
                        columns.append(col)
            for inl in variant.get("inlined_objects", []):
                for col_spec in inl["columns"]:
                    if col_spec["column"] not in columns:
                        columns.append(col_spec["column"])

    # Multivalued
    for mv in rules.get("multivalued", []):
        if mv.get("strategy") == "expand_rows":
            for item_spec in mv.get("per_item", []):
                if item_spec.get("type") == "quantity_value":
                    prefix = item_spec["prefix"]
                    columns.extend([f"{prefix}_value", f"{prefix}_unit"])
                else:
                    columns.append(item_spec["column"])
        elif mv.get("strategy") == "delimited":
            col_name = mv["path"].replace(".", "_")
            columns.append(col_name)

    return columns


def read_excel(
    excel_path: str | Path,
    sheet_names: list[str] | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """Read flat rows from an Excel workbook.

    Args:
        excel_path: Path to the .xlsx file
        sheet_names: Optional list of sheet names to read. If None, reads all.

    Returns:
        Dict mapping sheet name to list of flat row dicts
    """
    wb = load_workbook(str(excel_path), data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for ws in wb.worksheets:
        if sheet_names and ws.title not in sheet_names:
            continue

        rows_data: list[dict[str, Any]] = []
        headers: list[str] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(row)]
                continue

            # Skip empty rows
            if all(v is None for v in row):
                continue

            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    row_dict[headers[col_idx]] = value
            rows_data.append(row_dict)

        result[ws.title] = rows_data

    return result
